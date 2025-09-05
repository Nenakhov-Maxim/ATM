import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill

def create_empty_excel(columns: list, filename: str, sheet_name: str = 'Акта'):
    df = pd.DataFrame(columns=columns)

    if not os.path.exists('excel_files'):
        os.makedirs('excel_files')

    filepath = os.path.join('excel_files', filename)
    excel_writer = pd.ExcelWriter(filepath, engine='xlsxwriter')
    df.to_excel(excel_writer, index=False, sheet_name=sheet_name, freeze_panes=(1, 0))
    excel_writer._save()

    return filepath
  
def create_excel_from_dict_list(header_list: list, dict_list, output_filename: str, sheet_name='Sheet1'):
    # Создаем директорию, если она не существует
    if not os.path.exists('excel_files'):
        os.makedirs('excel_files')

    filepath = os.path.join('excel_files', output_filename)

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    ws.append(['Дата', '__________', '', 'Профильная линия', '', '', 'Смена', '___________' ])
    ws.append([])

    # Записываем данные из списка словарей в Excel
    ws.append(header_list)
    if dict_list:
        for key in dict_list:
            ws.append([dict_list[key]['label']])
            for data in dict_list[key]['data']:
                ws.append(data)

    # Настраиваем стили для красивого вида
    header_style = NamedStyle(name='header')
    header_style.font = Font(bold=True, color='FFFFFF')
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    border_style = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    header_style.border = border_style

    cell_style = NamedStyle(name='cell')
    cell_style.alignment = Alignment(horizontal='left', vertical='center')
    cell_style.border = border_style
    top_style = NamedStyle(name='top')
    top_style.alignment = Alignment(horizontal='right', vertical='center')
    
    for cell in ws[3]:  # Применяем стиль к заголовкам
        cell.style = header_style
    
    for cell in ws[1]:
      if cell.col_idx != 5:
        cell.style = top_style

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.style = cell_style

    ws.append([])
    ws.append(['Мастер','______________________'])
    # Автоматическое изменение ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем файл
    wb.save(filepath)
    return filepath
  
def create_tabel_users():
    filepath = create_excel_from_dict_list(dict_list=[{'Ф.И.О': 'Рабочий 1', 'Номер линии': '1', 'Марка изделия':'Т-профиль', 'Общее кол-во п/м':'2800',
                                                       'Отработанные часы':'7', 'Ср. ед.':'0', 'Хоз. работы':'Да', 'Подпись работника':''},
                                                      {'Ф.И.О': 'Рабочий 2', 'Номер линии': '1', 'Марка изделия':'Т-профиль', 'Общее кол-во п/м':'2800',
                                                       'Отработанные часы':'7', 'Ср. ед.':'0', 'Хоз. работы':'Да', 'Подпись работника':''}], output_filename='Акт №1.xlsx')
    #print(filepath)
